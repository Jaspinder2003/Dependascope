"""
stage1_inspect.py – Inspect all dataset files and emit schema reports.
Reads JSON files, XLSX files; never loads the full dataset into RAM.
"""

import argparse
import json
import logging
import os
import sys
import zipfile
from pathlib import Path
from typing import Any

import openpyxl

# Allow running from the pipeline/ dir
sys.path.insert(0, str(Path(__file__).parent))
from config import (
    DATA_DIR, PARTITION1, PARTITION2, DERIVED, OUTPUT_DIR, LOG_DIR,
    DATASET_SOURCES
)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler(LOG_DIR / "stage1_inspect.log", encoding="utf-8"),
    ]
)
logger = logging.getLogger(__name__)


def inspect_json_file(path: Path, max_items: int = 3) -> dict:
    """Inspect a single JSON data file – returns summary dict."""
    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)

    total_count = data.get("total_count", "?")
    items = data.get("items", [])
    num_items = len(items)

    # Collect all keys seen across the first few items
    all_keys: set = set()
    for item in items[:max_items]:
        all_keys.update(item.keys())

    # Grab representative record
    representative = None
    if items:
        representative = {
            k: (v if not isinstance(v, dict) else {kk: vv for kk, vv in list(v.items())[:5]})
            for k, v in list(items[0].items())[:10]
        }

    return {
        "file": str(path),
        "total_count": total_count,
        "items_in_file": num_items,
        "keys": sorted(all_keys),
        "representative_record": representative,
    }


def inspect_json_directory(dirpath: Path) -> dict:
    """Walk all JSON files under a directory; stream through them."""
    json_files = sorted(dirpath.rglob("*.json"))
    logger.info(f"Found {len(json_files)} JSON files under {dirpath}")

    total_items = 0
    all_keys: set = set()
    first_record = None
    file_summaries = []

    for i, jf in enumerate(json_files):
        try:
            with open(jf, "r", encoding="utf-8") as f:
                data = json.load(f)
            items = data.get("items", [])
            total_items += len(items)
            for item in items[:2]:
                all_keys.update(item.keys())
            if first_record is None and items:
                first_record = items[0]
            file_summaries.append({
                "file": jf.name,
                "size_bytes": jf.stat().st_size,
                "items_in_file": len(items),
                "total_count": data.get("total_count"),
            })
        except Exception as e:
            logger.warning(f"Error reading {jf}: {e}")
            file_summaries.append({"file": jf.name, "error": str(e)})

        if i % 20 == 0:
            logger.info(f"  Processed {i+1}/{len(json_files)} files, {total_items} items so far...")

    return {
        "directory": str(dirpath),
        "json_file_count": len(json_files),
        "total_items_across_all_files": total_items,
        "keys_observed": sorted(all_keys),
        "file_summaries": file_summaries,
        "first_record_sample": first_record,
    }


def inspect_zip_file(zippath: Path) -> dict:
    """List contents of a zip without extracting."""
    with zipfile.ZipFile(zippath, "r") as zf:
        names = zf.namelist()
        json_names = [n for n in names if n.endswith(".json")]
        # Peek at first JSON
        sample = None
        if json_names:
            with zf.open(json_names[0]) as f:
                try:
                    data = json.load(f)
                    items = data.get("items", [])
                    sample = {
                        "total_count": data.get("total_count"),
                        "items_in_file": len(items),
                        "keys": sorted(items[0].keys()) if items else [],
                    }
                except Exception as e:
                    sample = {"error": str(e)}
    return {
        "zip": str(zippath),
        "total_entries": len(names),
        "json_file_count": len(json_names),
        "first_json_sample": sample,
    }


def inspect_xlsx_file(path: Path, max_rows_preview: int = 3) -> dict:
    """Inspect an XLSX file – stream through for row count."""
    logger.info(f"Inspecting {path.name} ...")
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    result = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows_iter = ws.iter_rows(values_only=True)

        # Headers
        try:
            headers = list(next(rows_iter))
        except StopIteration:
            result[sheet_name] = {"headers": [], "row_count": 0}
            continue

        # Preview rows
        preview = []
        for i, row in enumerate(rows_iter):
            if i < max_rows_preview:
                preview.append(dict(zip(headers, [str(v)[:120] if v is not None else None for v in row])))

        row_count = ws.max_row - 1  # subtract header
        result[sheet_name] = {
            "headers": headers,
            "estimated_row_count": row_count,
            "preview_rows": preview,
        }

    wb.close()
    return {"file": str(path), "sheets": result}


def main(args: argparse.Namespace) -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    LOG_DIR.mkdir(parents=True, exist_ok=True)

    schema_report: dict[str, Any] = {
        "summary": {},
        "sources": {}
    }

    # ── 1. Partition 2 JSON directory (main Dependabot JSON data) ────────────
    logger.info("=== Partition 2 Dependabot JSON directory ===")
    json_dir = PARTITION2 / "Dataset" / "Dependabot"
    if json_dir.exists():
        json_dir_report = inspect_json_directory(json_dir)
        schema_report["sources"]["partition2_dependabot_json"] = json_dir_report
        logger.info(f"  Total JSON items: {json_dir_report['total_items_across_all_files']}")
        logger.info(f"  Keys: {json_dir_report['keys_observed']}")
    else:
        logger.warning(f"JSON dir not found: {json_dir}")

    # ── 2. Partition 2 ZIP files ──────────────────────────────────────────────
    for zipname in ["Dependabot.zip", "Dependabot-preview.zip"]:
        zippath = PARTITION2 / "Dataset" / zipname
        if zippath.exists():
            logger.info(f"=== {zipname} ===")
            zr = inspect_zip_file(zippath)
            schema_report["sources"][zipname] = zr
            logger.info(f"  JSON files: {zr['json_file_count']}, entries: {zr['total_entries']}")

    # ── 3. XLSX files ─────────────────────────────────────────────────────────
    xlsx_files = {
        "partition1_PRs": PARTITION1 / "PRs.xlsx",
        "partition1_Repos": PARTITION1 / "Repos.xlsx",
        "partition2_PRs": PARTITION2 / "PRs.xlsx",
        "partition2_Repos": PARTITION2 / "Repos.xlsx",
        "derived_subsample": DERIVED / "2. Extracted sub-sample.xlsx",
    }

    for label, path in xlsx_files.items():
        if path.exists():
            logger.info(f"=== {label}: {path.name} ===")
            try:
                xr = inspect_xlsx_file(path)
                schema_report["sources"][label] = xr
                for sheet, info in xr["sheets"].items():
                    logger.info(f"  Sheet '{sheet}': ~{info['estimated_row_count']} rows")
                    logger.info(f"  Headers: {info['headers']}")
            except Exception as e:
                logger.error(f"Error inspecting {path}: {e}")
                schema_report["sources"][label] = {"error": str(e)}
        else:
            logger.warning(f"File not found: {path}")

    # ── 4. Partition 1 RAR file ───────────────────────────────────────────────
    rar_path = PARTITION1 / "Dataset" / "Part 0.rar"
    schema_report["sources"]["partition1_rar"] = {
        "path": str(rar_path),
        "exists": rar_path.exists(),
        "size_bytes": rar_path.stat().st_size if rar_path.exists() else 0,
        "note": "RAR archive – requires 'unrar' or 7-Zip to extract. "
                "Contents likely mirror the Partition 2 JSON structure. "
                "Skip if Partition 2 JSON data is sufficient."
    }
    logger.info(f"RAR: {rar_path} ({'found' if rar_path.exists() else 'NOT FOUND'})")

    # ── 5. Summarise ──────────────────────────────────────────────────────────
    p2_json = schema_report["sources"].get("partition2_dependabot_json", {})
    p1_prs = schema_report["sources"].get("partition1_PRs", {})
    p2_prs = schema_report["sources"].get("partition2_PRs", {})

    p1_prs_rows = 0
    if p1_prs.get("sheets"):
        sheet = next(iter(p1_prs["sheets"].values()))
        p1_prs_rows = sheet.get("estimated_row_count", 0)

    p2_prs_rows = 0
    if p2_prs.get("sheets"):
        sheet = next(iter(p2_prs["sheets"].values()))
        p2_prs_rows = sheet.get("estimated_row_count", 0)

    schema_report["summary"] = {
        "partition2_json_dependabot_items": p2_json.get("total_items_across_all_files", 0),
        "partition2_json_file_count": p2_json.get("json_file_count", 0),
        "partition1_prs_xlsx_rows": p1_prs_rows,
        "partition2_prs_xlsx_rows": p2_prs_rows,
        "preferred_primary_source": "partition1_PRs (Excel) + partition2_dependabot_json (JSON)",
        "key_fields_available": [
            "Owner", "Repo", "Number", "Author", "Title", "Body",
            "State", "Created_at", "Merged_at", "Closed_at",
            "Comments", "Commits", "Additions", "Deletions", "Changed_files",
            "Labels", "Language", "Archived", "Mergeable", "Mergeable_state",
        ],
        "missing_from_local_data": [
            "head_sha", "before_sha (parent of head)",
            "merge_commit_sha",
            "changed_file_list_with_patches",
            "commit_list",
            "check_run_results",
        ],
        "github_enrichment_required": True,
    }

    # ── 6. Write reports ──────────────────────────────────────────────────────
    json_report_path = OUTPUT_DIR / "dataset_schema_report.json"
    with open(json_report_path, "w", encoding="utf-8") as f:
        json.dump(schema_report, f, indent=2, default=str)
    logger.info(f"JSON schema report written to {json_report_path}")

    # Human-readable markdown summary
    _write_markdown_report(schema_report, OUTPUT_DIR / "dataset_schema_report.md")
    logger.info(f"Markdown report written to {OUTPUT_DIR / 'dataset_schema_report.md'}")

    print("\nStage 1 complete. Reports written to:", OUTPUT_DIR)
    return schema_report


def _write_markdown_report(report: dict, path: Path) -> None:
    lines = [
        "# Dataset Schema Report",
        "",
        "## Summary",
        "",
    ]
    s = report.get("summary", {})
    lines += [
        f"| Key | Value |",
        f"|-----|-------|",
        f"| Partition 2 JSON items | {s.get('partition2_json_dependabot_items', 'N/A')} |",
        f"| Partition 2 JSON files | {s.get('partition2_json_file_count', 'N/A')} |",
        f"| Partition 1 PRs.xlsx rows | {s.get('partition1_prs_xlsx_rows', 'N/A')} |",
        f"| Partition 2 PRs.xlsx rows | {s.get('partition2_prs_xlsx_rows', 'N/A')} |",
        f"| GitHub enrichment required | {s.get('github_enrichment_required', True)} |",
        "",
        "## Available Fields (local data)",
        "",
        "```",
        "\n".join(s.get("key_fields_available", [])),
        "```",
        "",
        "## Fields Requiring GitHub API",
        "",
        "```",
        "\n".join(s.get("missing_from_local_data", [])),
        "```",
        "",
        "## Data Sources",
        "",
    ]

    for src_name, src_data in report.get("sources", {}).items():
        lines.append(f"### {src_name}")
        if "sheets" in src_data:
            for sheet, info in src_data["sheets"].items():
                lines.append(f"- Sheet: `{sheet}`")
                lines.append(f"  - Rows: ~{info.get('estimated_row_count', '?')}")
                lines.append(f"  - Headers: `{info.get('headers', [])}`")
        elif "total_items_across_all_files" in src_data:
            lines.append(f"- JSON files: {src_data.get('json_file_count')}")
            lines.append(f"- Total items: {src_data.get('total_items_across_all_files')}")
            lines.append(f"- Keys: `{src_data.get('keys_observed', [])}`")
        elif "zip" in src_data:
            lines.append(f"- ZIP entries: {src_data.get('total_entries')}")
            lines.append(f"- JSON files: {src_data.get('json_file_count')}")
        lines.append("")

    path.write_text("\n".join(lines), encoding="utf-8")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Stage 1: Inspect dataset files")
    parser.add_argument("--dry-run", action="store_true", help="Do not write output files")
    args = parser.parse_args()
    main(args)
