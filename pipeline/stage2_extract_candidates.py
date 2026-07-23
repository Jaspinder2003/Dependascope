"""
stage2_extract_candidates.py
────────────────────────────
Read all available data sources and build a deduplicated candidate table of
Dependabot PRs in the SQLite database + Parquet exports.

Sources consumed (in priority order):
  1. Partition 1 PRs.xlsx   – 363 k rows, fully structured (primary)
  2. Partition 2 PRs.xlsx   – ~46 k rows
  3. Partition 2 JSON files – raw GitHub Issues API format
  4. Derived subsample

All sources are streamed; no full file is loaded into RAM at once.
"""

import argparse
import json
import logging
import re
import sys
from pathlib import Path
from typing import Iterator, Optional
import sqlite3

import openpyxl

sys.path.insert(0, str(Path(__file__).parent))
import config as C
import db

LOG_DIR = C.LOG_DIR
LOG_DIR.mkdir(parents=True, exist_ok=True)
C.OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s – %(message)s",
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler(LOG_DIR / "stage2_extract.log", encoding="utf-8"),
    ]
)
logger = logging.getLogger(__name__)

# ─── Normalise author string ──────────────────────────────────────────────────
_DEPBOT_RE = re.compile(r"dependabot(?:-preview)?(?:\[bot\])?", re.I)


def is_dependabot_author(author: Optional[str]) -> bool:
    if not author:
        return False
    return bool(_DEPBOT_RE.fullmatch(author.strip()))


# ─── Title-based signals ──────────────────────────────────────────────────────
_BUMP_RE    = re.compile(r"(?i)(bump|update|upgrade)\s+\S+")
_GROUPED_RE = re.compile(r"(?i)(group|multiple|several|\d+\s+update|\d+\s+package)")
_SECURITY_RE = re.compile(r"(?i)\[security\]")


def classify_title(title: str) -> dict:
    """Return signals derived from the PR title alone."""
    title = title or ""
    return {
        "looks_like_bump": bool(_BUMP_RE.search(title)),
        "looks_grouped":   bool(_GROUPED_RE.search(title)),
        "is_security":     bool(_SECURITY_RE.search(title)),
    }


def extract_dep_from_title(title: str) -> tuple[Optional[str], Optional[str], Optional[str]]:
    """Best-effort extraction of (dep_name, old_version, new_version) from title."""
    title = title or ""
    # "Bump X from A to B"
    m = re.search(
        r"(?i)(?:bump|update|upgrade)\s+(\S+)\s+from\s+(\S+)\s+to\s+(\S+)",
        title
    )
    if m:
        return m.group(1), m.group(2), m.group(3)
    # "Bump X to B"
    m = re.search(r"(?i)(?:bump|update|upgrade)\s+(\S+)\s+to\s+(\S+)", title)
    if m:
        return m.group(1), None, m.group(2)
    return None, None, None


def classify_version_bump(old: Optional[str], new: Optional[str]) -> str:
    """Classify patch/minor/major/unknown from version strings."""
    if not old or not new:
        return "unknown"
    # Strip leading 'v'
    old_c = old.lstrip("v")
    new_c = new.lstrip("v")
    try:
        old_parts = [int(x) for x in old_c.split(".")[:3]]
        new_parts = [int(x) for x in new_c.split(".")[:3]]
        # Pad
        while len(old_parts) < 3: old_parts.append(0)
        while len(new_parts) < 3: new_parts.append(0)
        if new_parts[0] != old_parts[0]:
            return "major"
        if new_parts[1] != old_parts[1]:
            return "minor"
        if new_parts[2] != old_parts[2]:
            return "patch"
        return "unknown"
    except Exception:
        # SHA digest updates
        if re.match(r"[0-9a-f]{7,}", old_c) and re.match(r"[0-9a-f]{7,}", new_c):
            return "digest"
        return "unknown"


def pr_row_from_xlsx(headers: list, values: tuple, source_file: str, source_dataset: str) -> Optional[dict]:
    """Convert a raw xlsx row into a normalised PR dict. Returns None if clearly not Dependabot."""
    row = dict(zip(headers, values))

    author = str(row.get("Author") or "").strip()
    if not is_dependabot_author(author):
        return None

    title = str(row.get("Title") or "").strip()
    signals = classify_title(title)

    owner = str(row.get("Owner") or "").strip()
    repo  = str(row.get("Repo") or "").strip()
    if not owner or not repo:
        return None

    repo_full = f"{owner}/{repo}"
    pr_num    = row.get("Number")
    if pr_num is None:
        return None
    pr_num = int(pr_num)

    state     = str(row.get("State") or "").lower()
    labels_raw = str(row.get("Labels") or "")
    # Labels stored as string like "['dependencies', 'security']"
    try:
        labels_list = json.dumps(json.loads(labels_raw.replace("'", '"')))
    except Exception:
        labels_list = json.dumps([l.strip().strip("'\"[] ") for l in labels_raw.split(",")])

    dep, old_v, new_v = extract_dep_from_title(title)
    bump_type = classify_version_bump(old_v, new_v)

    strict_or_complex = "UNKNOWN"
    reason = ""
    if signals["looks_grouped"]:
        strict_or_complex = "COMPLEX"
        reason = "grouped_title_pattern"

    return {
        "repo":             repo_full,
        "pr_number":        pr_num,
        "pr_url":           f"https://github.com/{repo_full}/pull/{pr_num}",
        "author":           author,
        "title":            title,
        "body":             None,   # xlsx truncates body to int in some rows
        "state":            state,
        "created_at":       str(row.get("Created_at") or ""),
        "merged_at":        str(row.get("Merged_at") or ""),
        "closed_at":        str(row.get("Closed_at") or ""),
        "head_sha":         None,
        "before_sha":       None,
        "merge_sha":        None,
        "ecosystem":        _guess_ecosystem_from_language(
            str(row.get("Repos.Language") or row.get("Repo.Language") or "")
        ),
        "labels":           labels_list,
        "comments_count":   _safe_int(row.get("Comments")),
        "commits_count":    _safe_int(row.get("Commits")),
        "additions":        _safe_int(row.get("Additions")),
        "deletions":        _safe_int(row.get("Deletions")),
        "changed_files":    _safe_int(row.get("Changed_files")),
        "priority_score":   None,
        "strict_or_complex": strict_or_complex,
        "classification_reason": reason,
        "processing_status": "PENDING",
        "source_dataset":   source_dataset,
        "source_file":      source_file,
    }


def pr_row_from_json_item(item: dict, source_file: str) -> Optional[dict]:
    """Convert a GitHub Issues API item dict into a normalised PR dict."""
    user = item.get("user") or {}
    author = user.get("login", "")
    if not is_dependabot_author(author):
        return None

    # Must have pull_request field to be a PR (not issue)
    if "pull_request" not in item:
        return None

    repo_url = item.get("repository_url", "")
    # e.g. https://api.github.com/repos/owner/repo
    m = re.search(r"repos/([^/]+/[^/]+)$", repo_url)
    if not m:
        return None
    repo_full = m.group(1)

    title    = str(item.get("title") or "").strip()
    signals  = classify_title(title)
    pr_num   = item.get("number")
    if pr_num is None:
        return None

    state = str(item.get("state") or "").lower()
    labels_raw = [l.get("name", "") for l in (item.get("labels") or [])]
    labels_json = json.dumps(labels_raw)

    pr_data  = item.get("pull_request") or {}
    merged_at = pr_data.get("merged_at")

    dep, old_v, new_v = extract_dep_from_title(title)

    strict_or_complex = "UNKNOWN"
    reason = ""
    if signals["looks_grouped"]:
        strict_or_complex = "COMPLEX"
        reason = "grouped_title_pattern"

    return {
        "repo":              repo_full,
        "pr_number":         int(pr_num),
        "pr_url":            item.get("html_url", f"https://github.com/{repo_full}/pull/{pr_num}"),
        "author":            author,
        "title":             title,
        "body":              (item.get("body") or "")[:4000],
        "state":             state,
        "created_at":        str(item.get("created_at") or ""),
        "merged_at":         str(merged_at or ""),
        "closed_at":         str(item.get("closed_at") or ""),
        "head_sha":          None,
        "before_sha":        None,
        "merge_sha":         None,
        "ecosystem":         None,
        "labels":            labels_json,
        "comments_count":    _safe_int(item.get("comments")),
        "commits_count":     None,
        "additions":         None,
        "deletions":         None,
        "changed_files":     None,
        "priority_score":    None,
        "strict_or_complex": strict_or_complex,
        "classification_reason": reason,
        "processing_status": "PENDING",
        "source_dataset":    "partition2_json",
        "source_file":       source_file,
    }


def _safe_int(v) -> Optional[int]:
    try:
        return int(v)
    except Exception:
        return None


_LANG_ECO = {
    "javascript": "npm", "typescript": "npm", "vue": "npm",
    "python": "pip",
    "java": "maven",
    "kotlin": "gradle",
    "ruby": "gem",
    "go": "go",
    "rust": "cargo",
    "php": "composer",
    "c#": "nuget", "f#": "nuget",
}


def _guess_ecosystem_from_language(lang: str) -> Optional[str]:
    return _LANG_ECO.get(lang.lower(), None)


# ─── Streaming readers ────────────────────────────────────────────────────────

def stream_xlsx(path: Path, max_rows: Optional[int] = None) -> Iterator[tuple]:
    """Yield (headers, row_values) tuples from all sheets without loading full file."""
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    count = 0
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = ws.iter_rows(values_only=True)
        try:
            headers = [str(h) if h is not None else "" for h in next(rows)]
        except StopIteration:
            continue
        for row in rows:
            yield headers, row
            count += 1
            if max_rows and count >= max_rows:
                wb.close()
                return
    wb.close()


def stream_json_dir(dirpath: Path, max_items: Optional[int] = None) -> Iterator[dict]:
    """Yield individual JSON PR items from all files in a directory, streaming."""
    json_files = sorted(dirpath.rglob("*.json"))
    count = 0
    for jf in json_files:
        try:
            with open(jf, "r", encoding="utf-8") as f:
                data = json.load(f)
            for item in data.get("items", []):
                yield item
                count += 1
                if max_items and count >= max_items:
                    return
        except Exception as e:
            logger.warning(f"Skipping {jf}: {e}")


# ─── Main ─────────────────────────────────────────────────────────────────────

def main(args: argparse.Namespace) -> None:
    conn = db.init_db(C.DB_PATH)
    inserted = 0
    skipped  = 0
    total    = 0

    limit = args.limit

    # ── Source 1: Partition 1 PRs.xlsx ────────────────────────────────────────
    p1_prs = C.PARTITION1 / "PRs.xlsx"
    if p1_prs.exists():
        logger.info(f"Reading {p1_prs.name} ...")
        for i, (headers, row) in enumerate(stream_xlsx(p1_prs, max_rows=limit)):
            total += 1
            pr = pr_row_from_xlsx(headers, row, str(p1_prs), "partition1_xlsx")
            if pr:
                try:
                    db.upsert_pr(conn, pr)
                    inserted += 1
                except Exception as e:
                    logger.debug(f"Insert error row {i}: {e}")
                    skipped += 1
            else:
                skipped += 1

            if total % 50000 == 0:
                conn.commit()
                logger.info(f"  P1 XLSX: {total} rows read, {inserted} inserted, {skipped} skipped")

            if limit and inserted >= limit:
                break
        conn.commit()
        logger.info(f"Partition 1 XLSX done: {total} rows, {inserted} dependabot PRs")

    # ── Source 2: Partition 2 PRs.xlsx ────────────────────────────────────────
    p2_prs = C.PARTITION2 / "PRs.xlsx"
    if p2_prs.exists() and (not limit or inserted < limit):
        logger.info(f"Reading {p2_prs.name} ...")
        p2_inserted = 0
        for headers, row in stream_xlsx(p2_prs, max_rows=limit):
            pr = pr_row_from_xlsx(headers, row, str(p2_prs), "partition2_xlsx")
            if pr:
                try:
                    db.upsert_pr(conn, pr)
                    p2_inserted += 1
                    inserted += 1
                except Exception as e:
                    logger.debug(f"P2 XLSX insert error: {e}")
        conn.commit()
        logger.info(f"Partition 2 XLSX done: {p2_inserted} new dependabot PRs")

    # ── Source 3: Partition 2 JSON directory ──────────────────────────────────
    json_dir = C.PARTITION2 / "Dataset" / "Dependabot"
    if json_dir.exists() and (not limit or inserted < limit):
        logger.info("Reading Partition 2 JSON directory ...")
        json_inserted = 0
        json_total    = 0
        for item in stream_json_dir(json_dir, max_items=limit):
            json_total += 1
            # Determine source_file from path (approximate)
            pr = pr_row_from_json_item(item, source_file="partition2_json_dir")
            if pr:
                try:
                    db.upsert_pr(conn, pr)
                    json_inserted += 1
                    inserted += 1
                except Exception as e:
                    logger.debug(f"JSON insert error: {e}")
            if json_total % 2000 == 0:
                conn.commit()
                logger.info(f"  JSON: {json_total} items, {json_inserted} inserted")
            if limit and inserted >= limit:
                break
        conn.commit()
        logger.info(f"JSON dir done: {json_total} items, {json_inserted} dependabot PRs")

    # ── Totals ────────────────────────────────────────────────────────────────
    total_in_db = conn.execute("SELECT COUNT(*) FROM pull_requests").fetchone()[0]
    complex_count = conn.execute(
        "SELECT COUNT(*) FROM pull_requests WHERE strict_or_complex='COMPLEX'"
    ).fetchone()[0]
    unknown_count = conn.execute(
        "SELECT COUNT(*) FROM pull_requests WHERE strict_or_complex='UNKNOWN'"
    ).fetchone()[0]

    logger.info(f"Total PRs in DB: {total_in_db}")
    logger.info(f"  COMPLEX (grouped title): {complex_count}")
    logger.info(f"  UNKNOWN (not yet validated): {unknown_count}")

    # ── Export to Parquet ─────────────────────────────────────────────────────
    if not args.dry_run:
        _export_parquet(conn)

    conn.close()
    print(f"\nStage 2 complete. {total_in_db} candidate PRs in {C.DB_PATH}")


def _export_parquet(conn: sqlite3.Connection) -> None:
    """Export candidate table to parquet via pyarrow."""
    try:
        import pyarrow as pa
        import pyarrow.parquet as pq

        rows = conn.execute("SELECT * FROM pull_requests").fetchall()
        if not rows:
            return

        columns = [d[0] for d in conn.execute("SELECT * FROM pull_requests LIMIT 0").description]
        data = {col: [] for col in columns}
        for row in rows:
            for col, val in zip(columns, row):
                data[col].append(val)

        table = pa.table(data)
        out = C.OUTPUT_DIR / "dependabot_candidates.parquet"
        pq.write_table(table, str(out))
        logger.info(f"Parquet exported: {out}")
    except Exception as e:
        logger.warning(f"Parquet export failed: {e}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Stage 2: Extract Dependabot candidate PRs")
    parser.add_argument("--limit", type=int, default=None,
                        help="Max rows to process (for testing)")
    parser.add_argument("--dry-run", action="store_true",
                        help="Skip Parquet export")
    parser.add_argument("--resume", action="store_true",
                        help="(no-op in this stage; DB upsert handles resumption)")
    args = parser.parse_args()
    main(args)
